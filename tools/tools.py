from openpyxl import load_workbook
from typing import List, Any, Dict
import random
from langchain.tools import tool
from core.logger import setup_logger
from tools.utils import get_detailed_data_types

logger = setup_logger(__name__)


# @tool
# def get_sheet_names(file_path: str) -> List[str]:
#     """Get all sheet names from the Excel file."""
#     logger.info("Getting sheet names from %s", file_path)
#     workbook = load_workbook(file_path)
#     result = workbook.sheetnames
#     logger.info("Found %d sheets: %s", len(result), result)
#     return result


@tool
def get_row_values(file_path: str, sheet_name: str, row_number: int) -> List[Dict[str, Any]]:
    """Get all values from a specific row in the Excel sheet with their cell references."""
    logger.info("Getting row %d values from sheet '%s' in %s", row_number, sheet_name, file_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    result = []
    for cell in sheet[row_number]:
        if cell.value is not None:  # Only include non-empty cells
            result.append({
                "cell_reference": cell.coordinate,
                "value": cell.value
            })
    logger.info("Row %d has %d non-empty values", row_number, len(result))
    return result


@tool
def get_column_values(file_path: str, sheet_name: str, column_letter: str) -> List[Dict[str, Any]]:
    """Get all values from a specific column in the Excel sheet with their cell references."""
    logger.info("Getting column %s values from sheet '%s' in %s", column_letter, sheet_name, file_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    result = []
    for cell in sheet[column_letter]:
        if cell.value is not None:  # Only include non-empty cells
            result.append({
                "cell_reference": cell.coordinate,
                "value": cell.value
            })
    logger.info("Column %s has %d non-empty values", column_letter, len(result))
    return result


@tool
def get_cell_value(file_path: str, sheet_name: str, cell_reference: str) -> Any:
    """Get the value of a specific cell in the Excel sheet."""
    logger.info("Getting cell %s value from sheet '%s' in %s", cell_reference, sheet_name, file_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    result = sheet[cell_reference].value
    logger.info("Cell %s value: %s", cell_reference, result)
    return result


@tool
def get_data_types_column(file_path: str, sheet_name: str, column_letter: str) -> List[str]:
    """Get the data types of all values in a specific column."""
    logger.info("Getting data types for column %s in sheet '%s' from %s", column_letter, sheet_name, file_path)
    # Read the column values directly instead of calling get_column_values
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    result = [cell.value for cell in sheet[column_letter]]
    logger.info("Column %s has %d values", column_letter, len(result))
    return get_detailed_data_types(result)


@tool
def get_sheet_dimensions(file_path: str, sheet_name: str) -> Dict[str, int]:
    """Get the number of rows and columns in the Excel sheet."""
    logger.info("Getting dimensions for sheet '%s' in %s", sheet_name, file_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    result = {"rows": sheet.max_row, "columns": sheet.max_column}
    logger.info("Sheet '%s' dimensions: %d rows x %d columns", sheet_name, result['rows'], result['columns'])
    return result


@tool
def get_max_rows(file_path: str, sheet_name: str) -> int:
    """Get the maximum number of rows in the Excel sheet."""
    logger.info("Getting max rows for sheet '%s' in %s", sheet_name, file_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    result = sheet.max_row
    logger.info("Sheet '%s' has %d rows", sheet_name, result)
    return result


@tool
def get_max_columns(file_path: str, sheet_name: str) -> int:
    """Get the maximum number of columns in the Excel sheet."""
    logger.info("Getting max columns for sheet '%s' in %s", sheet_name, file_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    result = sheet.max_column
    logger.info("Sheet '%s' has %d columns", sheet_name, result)
    return result


# @tool
# def get_header_row(file_path: str, sheet_name: str, header_row: int = 1) -> List[str]:
#     """Get the header row values from the Excel sheet."""
#     logger.info(f"Getting header row {header_row} from sheet '{sheet_name}' in {file_path}")
#     result = get_row_values(file_path, sheet_name, header_row)
#     logger.info(f"Header row contains {len(result)} columns")
#     return result


@tool
def find_cells_with_value(file_path: str, sheet_name: str, search_value: Any) -> List[str]:
    """Find all cell references that contain a specific value."""
    logger.info("Searching for value '%s' in sheet '%s' from %s", search_value, sheet_name, file_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    cells = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == search_value:
                cells.append(cell.coordinate)
    logger.info("Found %d cells with value '%s': %s", len(cells), search_value, cells)
    return cells


@tool
def get_range_values(file_path: str, sheet_name: str, start_cell: str, end_cell: str) -> List[List[Any]]:
    """Get values from a range of cells in the Excel sheet."""
    logger.info("Getting range %s:%s from sheet '%s' in %s", start_cell, end_cell, sheet_name, file_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    range_data = sheet[start_cell:end_cell]
    result = [[cell.value for cell in row] for row in range_data]
    logger.info("Range %s:%s contains %d rows", start_cell, end_cell, len(result))
    return result


@tool
def get_sheet_content(file_path: str, sheet_name: str) -> Dict[int, Dict[str, Any]]:
    """Get all content of the sheet as a nested dictionary where outer dict keys are row numbers and inner dict keys are column letters."""
    logger.info("Getting all content from sheet '%s' in %s", sheet_name, file_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    result = {}
    for row in sheet.iter_rows():
        row_dict = {}
        for cell in row:
            if cell.value is not None:  # Only include non-empty cells
                row_dict[cell.column_letter] = cell.value
        if row_dict:  # Only include rows with data
            result[cell.row] = row_dict
    
    logger.info("Retrieved %d rows with data from sheet '%s'", len(result), sheet_name)
    return result


@tool
def get_sheet_content_sample(file_path: str, sheet_name: str, num_rows: int = 10, num_columns: int = 10) -> Dict[int, Dict[str, Any]]:
    """Get a sample of the sheet content by taking the first num_rows rows and num_columns columns."""
    logger.info("Getting sample content (%d rows x %d columns) from sheet '%s' in %s", num_rows, num_columns, sheet_name, file_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    # Get the actual dimensions of the sheet
    max_row = min(sheet.max_row, num_rows)
    max_col = min(sheet.max_column, num_columns)
    
    result = {}
    for row_num in range(1, max_row + 1):
        row_dict = {}
        for col_num in range(1, max_col + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            if cell.value is not None:  # Only include non-empty cells
                row_dict[cell.column_letter] = cell.value
        if row_dict:  # Only include rows with data
            result[row_num] = row_dict
    
    logger.info("Retrieved sample of %d rows with data from sheet '%s' (sampled %d rows x %d columns)", len(result), sheet_name, max_row, max_col)
    return result


@tool
def get_row_values_sample(file_path: str, sheet_name: str, row_number: int, sample_size: int = 10) -> List[Any]:
    """Get a random sample of values from a specific row in the Excel sheet."""
    logger.info("Getting sample of %d values from row %d in sheet '%s' from %s", sample_size, row_number, sheet_name, file_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    all_values = [cell.value for cell in sheet[row_number]]
    
    # Filter out None values and get non-empty values
    non_empty_values = [val for val in all_values if val is not None]
    
    if len(non_empty_values) <= sample_size:
        result = non_empty_values
        logger.info("Row %d has %d non-empty values (less than sample size)", row_number, len(result))
    else:
        result = random.sample(non_empty_values, sample_size)
        logger.info("Sampled %d values from row %d (total non-empty: %d)", len(result), row_number, len(non_empty_values))
    
    return result


@tool
def get_column_values_sample(file_path: str, sheet_name: str, column_letter: str, sample_size: int = 10) -> List[Any]:
    """Get a random sample of values from a specific column in the Excel sheet."""
    logger.info("Getting sample of %d values from column %s in sheet '%s' from %s", sample_size, column_letter, sheet_name, file_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    all_values = [cell.value for cell in sheet[column_letter]]
    
    # Filter out None values and get non-empty values
    non_empty_values = [val for val in all_values if val is not None]
    
    if len(non_empty_values) <= sample_size:
        result = non_empty_values
        logger.info("Column %s has %d non-empty values (less than sample size)", column_letter, len(result))
    else:
        result = random.sample(non_empty_values, sample_size)
        logger.info("Sampled %d values from column %s (total non-empty: %d)", len(result), column_letter, len(non_empty_values))
    
    return result


@tool
def get_data_types_column_sample(file_path: str, sheet_name: str, column_letter: str, sample_size: int = 10) -> List[str]:
    """Get the data types of a random sample of values from a specific column."""
    logger.info("Getting data types for sample of %d values from column %s in sheet '%s' from %s", sample_size, column_letter, sheet_name, file_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    all_values = [cell.value for cell in sheet[column_letter]]
    
    # Filter out None values and get non-empty values
    non_empty_values = [val for val in all_values if val is not None]
    
    if len(non_empty_values) <= sample_size:
        sample_values = non_empty_values
        logger.info("Column %s has %d non-empty values (less than sample size)", column_letter, len(sample_values))
    else:
        sample_values = random.sample(non_empty_values, sample_size)
        logger.info("Sampled %d values from column %s (total non-empty: %d)", len(sample_values), column_letter, len(non_empty_values))
    
    result = get_detailed_data_types(sample_values)
    return result


@tool
def get_nonempty_column_letters(file_path: str, sheet_name: str) -> List[str]:
    """Get a list of column letters that contain non-empty values in the Excel sheet."""
    logger.info("Getting non-empty column letters from sheet '%s' in %s", sheet_name, file_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    nonempty_columns = []
    for column in sheet.columns:
        # Check if any cell in this column has a non-empty value
        has_data = any(cell.value is not None for cell in column)
        if has_data:
            nonempty_columns.append(column[0].column_letter)
    
    logger.info("Found %d non-empty columns: %s", len(nonempty_columns), nonempty_columns)
    return nonempty_columns

