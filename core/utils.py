import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Border
from typing import List, Optional, Dict
from core.logger import setup_logger
from openpyxl import load_workbook

logger = setup_logger(__name__)

def remove_hidden_columns(file_path: str, sheet_name: str, output_path: Optional[str] = None) -> List[str]:
    """Remove hidden and grouped columns from an Excel sheet and return list of removed columns."""
    logger.info("Removing hidden columns from sheet '%s' in %s", sheet_name, file_path)
    
    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    logger.info("Successfully loaded workbook and sheet '%s'", sheet_name)
    
    # Get all column letters
    max_col = ws.max_column
    cols = [get_column_letter(i) for i in range(1, max_col + 1)]
    logger.info("Found %d total columns in sheet '%s'", len(cols), sheet_name)
    
    # Find hidden columns
    hidden_cols = []
    last_hidden = 0
    for i, col in enumerate(cols):
        # Column is hidden
        if ws.column_dimensions[col].hidden:
            hidden_cols.append(col)
            # Last column in the hidden group
            last_hidden = ws.column_dimensions[col].max
        # Appending column if more columns in the group
        elif i + 1 <= last_hidden:
            hidden_cols.append(col)
    
    logger.info("Found %d hidden columns in sheet '%s': %s", len(hidden_cols), sheet_name, hidden_cols)
    
    # Get visible columns
    visible_cols = [col for col in cols if col not in hidden_cols]
    logger.info("Remaining visible columns in sheet '%s': %s", sheet_name, visible_cols)
    
    # Clear hidden column data
    cols_to_delete = [column_index_from_string(col) for col in hidden_cols]
    for col_idx in sorted(set(cols_to_delete), reverse=True):
        for row in range(2, ws.max_row + 1):  # start from 2 to skip header
            cell = ws[f"{get_column_letter(col_idx)}{row}"]
            cell.value = None
            cell.border = Border()
    
    logger.info("Cleared data from %d hidden columns in sheet '%s'", len(cols_to_delete), sheet_name)
    
    # Reset column dimensions for hidden columns
    for col, dim in ws.column_dimensions.items():
        if ws.column_dimensions[col].hidden:
            ws.column_dimensions[col].hidden = False
            ws.column_dimensions[col].outlineLevel = False
            ws.column_dimensions[col].max = column_index_from_string(col)
    
    logger.info("Reset column dimensions for hidden columns in sheet '%s'", sheet_name)
    
    # Save to output path or original file
    if output_path:
        wb.save(output_path)
        logger.info("Saved cleaned sheet '%s' to %s", sheet_name, output_path)
    else:
        wb.save(file_path)
        logger.info("Saved cleaned sheet '%s' to original file %s", sheet_name, file_path)
    
    logger.info("Successfully removed %d hidden columns from sheet '%s'", len(hidden_cols), sheet_name)
    return hidden_cols


def remove_hidden_columns_all_sheets(file_path: str, output_path: Optional[str] = None) -> Dict[str, List[str]]:
    """Remove hidden and grouped columns from all sheets in an Excel workbook and return dict mapping sheet names to removed columns."""
    logger.info("Removing hidden columns from all sheets in %s", file_path)
    
    # Load workbook to get sheet names
    wb = openpyxl.load_workbook(file_path)
    sheet_names = wb.sheetnames
    wb.close()
    logger.info("Found %d sheets to process: %s", len(sheet_names), sheet_names)
    
    # Dictionary to store results for each sheet
    results = {}
    
    # Process each sheet by calling remove_hidden_columns
    for sheet_name in sheet_names:
        logger.info("Processing sheet '%s' (%d/%d)", sheet_name, sheet_names.index(sheet_name) + 1, len(sheet_names))
        # Call remove_hidden_columns for this specific sheet
        hidden_cols = remove_hidden_columns(file_path, sheet_name)
        results[sheet_name] = hidden_cols
        logger.info("Completed processing sheet '%s' - removed %d hidden columns", sheet_name, len(hidden_cols))
    
    total_removed = sum(len(cols) for cols in results.values())
    logger.info("Successfully processed all %d sheets, removed %d total hidden columns", len(sheet_names), total_removed)
    return results


def get_sheet_names(file_path: str) -> List[str]:
    """Get all sheet names from the Excel file."""
    logger.info("Getting sheet names from %s", file_path)
    workbook = load_workbook(file_path)
    result = workbook.sheetnames
    logger.info("Successfully loaded workbook with %d sheets: %s", len(result), result)
    workbook.close()
    return result
